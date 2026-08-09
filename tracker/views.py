from datetime import date, timedelta

from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Exists, OuterRef, Prefetch, Q
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from accounts.models import Department

from .decorators import role_required
from .forms import (
    TaskArtifactForm,
    TaskArtifactFormSet,
    TaskForm,
    TaskInlineEditForm,
    TaskPopupCreateForm,
    TaskStatusUpdateForm,
    TaskWeeklyStatusForm,
)
from .models import Task, TaskArtifact, TaskStatus, TaskWeeklyStatus
from .utils import get_week_start


STATUS_SUMMARY_GROUPINGS = {
    "department-assignee",
    "department",
    "assignee",
    "none",
}


def get_status_summary_grouping(user, requested_grouping):
    role_code = user.role.code

    if role_code == "employee":
        return "none"

    if role_code == "manager":
        if requested_grouping in {"assignee", "none"}:
            return requested_grouping
        return "assignee"

    if requested_grouping in STATUS_SUMMARY_GROUPINGS:
        return requested_grouping

    return "department-assignee"


def build_status_summary_groups(tasks, grouping):
    if grouping == "none":
        return [
            {
                "key": "all",
                "label": "",
                "tasks": tasks,
                "missing_count": sum(
                    not task.current_week_status_filled
                    for task in tasks
                ),
            }
        ]

    groups = []
    groups_by_key = {}

    for task in tasks:
        if grouping == "department-assignee":
            key = f"{task.department_id}:{task.assignee_id}"
            label = f"{task.department.name} / {task.assignee.name}"
        elif grouping == "department":
            key = f"department:{task.department_id}"
            label = task.department.name
        else:
            key = f"assignee:{task.assignee_id}"
            label = task.assignee.name

        group = groups_by_key.get(key)

        if group is None:
            group = {
                "key": key,
                "label": label,
                "tasks": [],
                "missing_count": 0,
            }
            groups_by_key[key] = group
            groups.append(group)

        group["tasks"].append(task)

        if not task.current_week_status_filled:
            group["missing_count"] += 1

    return groups


def build_status_summary_workbook(tasks, week_starts):
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Сводка по статусам"

    headers = [
        "Отдел",
        "Стрим",
        "Номер задачи",
        "Описание",
        "Артефакты",
        "Ответственный",
        "Статус задачи",
        *[
            f"Статус за неделю с {week_start.strftime('%d.%m.%Y')}"
            for week_start in week_starts
        ],
    ]
    worksheet.append(headers)

    header_fill = PatternFill("solid", fgColor="E5E7EB")

    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for task in tasks:
        artifacts = "\n".join(
            f"{artifact.name}: {artifact.url}"
            for artifact in task.summary_artifacts
        )
        row = [
            task.department.name,
            task.project_stream.name,
            task.external_number,
            task.summary,
            artifacts,
            task.assignee.name,
            task.status.name,
            *[status["text"] for status in task.summary_statuses],
        ]
        worksheet.append(row)
        row_number = worksheet.max_row

        if task.external_url and task.external_number:
            number_cell = worksheet.cell(row=row_number, column=3)
            number_cell.hyperlink = task.external_url
            number_cell.style = "Hyperlink"

        for cell in worksheet[row_number]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [18, 24, 18, 42, 32, 24, 20]
    widths.extend([38] * len(week_starts))

    for column_number, width in enumerate(widths, start=1):
        worksheet.column_dimensions[
            get_column_letter(column_number)
        ].width = width

    worksheet.freeze_panes = "H2"
    worksheet.auto_filter.ref = worksheet.dimensions

    return workbook


@login_required
def dashboard(request):
    return render(request, "tracker/dashboard.html")


@role_required("administrator")
def administration(request):
    return render(request, "tracker/administration.html")


@login_required
def task_list(request):
    current_week_start = get_week_start(timezone.localdate())
    latest_previous_week_start = current_week_start - timedelta(weeks=1)

    selected_week_value = request.GET.get("week")

    if selected_week_value:
        try:
            selected_date = date.fromisoformat(selected_week_value)
            previous_week_start = get_week_start(selected_date)
        except ValueError:
            previous_week_start = latest_previous_week_start
    else:
        previous_week_start = latest_previous_week_start

    if previous_week_start >= current_week_start:
        previous_week_start = latest_previous_week_start

    tasks = Task.objects.select_related(
        "project_stream",
        "status",
        "assignee",
        "created_by",
        "department",
    )

    if request.user.role.code == "employee":
        tasks = tasks.filter(assignee=request.user)

    elif request.user.role.code == "manager":
        if request.user.department_id:
            tasks = tasks.filter(
                department_id=request.user.department_id,
            )
        else:
            tasks = tasks.none()

    show_done = request.GET.get("show_done") == "1"
    show_cancelled = request.GET.get("show_cancelled") == "1"

    hidden_status_codes = []

    if not show_done:
        hidden_status_codes.append("done")

    if not show_cancelled:
        hidden_status_codes.append("cancelled")

    if hidden_status_codes:
        tasks = tasks.exclude(
            status__code__in=hidden_status_codes,
        )

    search_value = request.GET.get("search", "").strip()

    if search_value:
        tasks = tasks.filter(
            Q(external_number__icontains=search_value)
            | Q(summary__icontains=search_value)
            | Q(project_stream__name__icontains=search_value)
            | Q(assignee__name__icontains=search_value)
            | Q(department__code__icontains=search_value)
            | Q(department__name__icontains=search_value)
        )

    sort_fields = {
        "project": "project_stream__name",
        "number": "external_number",
        "summary": "summary",
    }

    sort_value = request.GET.get("sort", "project")
    sort_direction = request.GET.get("direction", "asc")

    if sort_value not in sort_fields:
        sort_value = "project"

    sort_field = sort_fields[sort_value]

    if sort_direction == "desc":
        sort_field = f"-{sort_field}"
    else:
        sort_direction = "asc"

    tasks = tasks.order_by(sort_field, "id")

    paginator = Paginator(tasks, 10)
    page_obj = paginator.get_page(request.GET.get("page"))
    tasks = list(page_obj.object_list)

    pagination_params = request.GET.copy()
    pagination_params.pop("page", None)
    pagination_query = pagination_params.urlencode()

    accessible_task_ids = [task.id for task in tasks]

    create_form = TaskPopupCreateForm(user=request.user)
    open_create_modal = False

    if request.method == "POST" and request.POST.get("action") == "create_task":
        create_form = TaskPopupCreateForm(
            request.POST,
            user=request.user,
        )
        open_create_modal = True

        if create_form.is_valid():
            create_form.save()
            return redirect(request.get_full_path())

        task_id = None
        form = None
    elif request.method == "POST":
        task_id = request.POST.get("task_id")

        try:
            task_id = int(task_id)
        except (TypeError, ValueError):
            raise PermissionDenied

        task = get_object_or_404(Task, id=task_id)
        check_task_access(request.user, task)

        weekly_status = TaskWeeklyStatus.objects.filter(
            task=task,
            week_start=current_week_start,
        ).first()

        form = TaskWeeklyStatusForm(
            request.POST,
            instance=weekly_status,
            prefix=f"task-{task.id}",
        )

        if form.is_valid():
            weekly_status = form.save(commit=False)
            weekly_status.task = task
            weekly_status.week_start = current_week_start
            weekly_status.updated_by = request.user
            weekly_status.save()

            return redirect(request.get_full_path())
    else:
        task_id = None
        form = None

    weekly_statuses = TaskWeeklyStatus.objects.filter(
        task_id__in=accessible_task_ids,
        week_start__in=[
            current_week_start,
            previous_week_start,
        ],
    )

    current_statuses = {}
    previous_statuses = {}

    for weekly_status in weekly_statuses:
        if weekly_status.week_start == current_week_start:
            current_statuses[weekly_status.task_id] = weekly_status
        elif weekly_status.week_start == previous_week_start:
            previous_statuses[weekly_status.task_id] = weekly_status

    for task in tasks:
        current_status = current_statuses.get(task.id)

        task.current_weekly_status = current_status
        task.previous_weekly_status = previous_statuses.get(task.id)

        if (
            request.method == "POST"
            and task.id == task_id
            and form is not None
        ):
            task.weekly_status_form = form
        else:
            task.weekly_status_form = TaskWeeklyStatusForm(
                instance=current_status,
                prefix=f"task-{task.id}",
            )

        task.inline_edit_form = TaskInlineEditForm(
            instance=task,
            prefix=f"task-edit-{task.id}",
            user=request.user,
        )

        task.status_update_form = TaskStatusUpdateForm(
            instance=task,
            prefix=f"task-status-{task.id}",
        )

        task.status_update_form.fields["status"].widget.attrs[
            "data-status-code"
        ] = task.status.code

    older_week_start = previous_week_start - timedelta(weeks=1)
    newer_week_start = previous_week_start + timedelta(weeks=1)

    can_move_forward = newer_week_start < current_week_start

    if request.GET.get("format") == "json":
        previous_status_data = {}

        for task in tasks:
            weekly_status = task.previous_weekly_status

            previous_status_data[str(task.id)] = {
                "text": weekly_status.text if weekly_status else "",
                "has_status": weekly_status is not None,
            }

        return JsonResponse(
            {
                "previous_week_start": previous_week_start.isoformat(),
                "previous_week_label": previous_week_start.strftime(
                    "%d.%m.%Y"
                ),
                "older_week_start": older_week_start.isoformat(),
                "newer_week_start": newer_week_start.isoformat(),
                "can_move_forward": can_move_forward,
                "statuses": previous_status_data,
            }
        )

    return render(
        request,
        "tracker/task_list.html",
        {
            "tasks": tasks,
            "current_week_start": current_week_start,
            "previous_week_start": previous_week_start,
            "older_week_start": older_week_start,
            "newer_week_start": newer_week_start,
            "can_move_forward": can_move_forward,
            "selected_sort": sort_value,
            "sort_direction": sort_direction,
            "selected_search": search_value,
            "show_done": show_done,
            "show_cancelled": show_cancelled,
            "page_obj": page_obj,
            "pagination_query": pagination_query,
            "create_form": create_form,
            "open_create_modal": open_create_modal,
        },
    )


@login_required
def status_summary(request):
    current_week_start = get_week_start(timezone.localdate())
    week_starts = [
        current_week_start - timedelta(weeks=offset)
        for offset in range(3, -1, -1)
    ]
    role_code = request.user.role.code

    current_status_exists = TaskWeeklyStatus.objects.filter(
        task_id=OuterRef("pk"),
        week_start=current_week_start,
    ).exclude(text="")

    tasks = (
        Task.objects.select_related(
            "project_stream",
            "status",
            "assignee",
            "department",
        )
        .annotate(
            current_week_status_filled=Exists(
                current_status_exists,
            )
        )
        .prefetch_related(
            Prefetch(
                "artifacts",
                queryset=TaskArtifact.objects.order_by("created_at"),
                to_attr="summary_artifacts",
            ),
            Prefetch(
                "weekly_statuses",
                queryset=TaskWeeklyStatus.objects.filter(
                    week_start__in=week_starts,
                ).order_by("week_start"),
                to_attr="summary_weekly_statuses",
            ),
        )
    )

    if role_code == "employee":
        tasks = tasks.filter(assignee=request.user)
    elif role_code == "manager":
        if request.user.department_id:
            tasks = tasks.filter(
                department_id=request.user.department_id,
            )
        else:
            tasks = tasks.none()
    elif role_code not in {"head", "administrator"}:
        tasks = tasks.none()

    selected_department = request.GET.get("department", "").strip()

    if role_code in {"head", "administrator"} and selected_department:
        try:
            selected_department_id = int(selected_department)
        except ValueError:
            selected_department = ""
        else:
            tasks = tasks.filter(
                department_id=selected_department_id,
            )
    else:
        selected_department = ""

    search_value = request.GET.get("search", "").strip()

    if search_value:
        tasks = tasks.filter(
            Q(external_number__icontains=search_value)
            | Q(summary__icontains=search_value)
            | Q(project_stream__name__icontains=search_value)
            | Q(assignee__name__icontains=search_value)
            | Q(department__name__icontains=search_value)
            | Q(department__code__icontains=search_value)
        )

    requested_show_final = request.GET.get("show_final") == "1"
    final_only = request.GET.get("final_only") == "1"

    if final_only:
        show_final_before_final_only = request.GET.get(
            "show_final_before_final_only",
            "0",
        )

        if show_final_before_final_only not in {"0", "1"}:
            show_final_before_final_only = "0"

        show_final = True
        tasks = tasks.filter(
            status__code__in=["done", "cancelled"],
        )
    else:
        show_final = requested_show_final
        show_final_before_final_only = (
            "1" if show_final else "0"
        )

    if not show_final:
        tasks = tasks.exclude(
            status__code__in=["done", "cancelled"],
        )

    missing_only = request.GET.get("missing_only") == "1"

    if missing_only:
        tasks = tasks.filter(current_week_status_filled=False)

    tasks = list(
        tasks.order_by(
            "department__name",
            "assignee__name",
            "project_stream__name",
            "external_number",
            "id",
        )
    )

    for task in tasks:
        statuses_by_week = {
            status.week_start: status
            for status in task.summary_weekly_statuses
        }
        task.summary_statuses = [
            {
                "week_start": week_start,
                "text": (
                    statuses_by_week[week_start].text
                    if week_start in statuses_by_week
                    else ""
                ),
            }
            for week_start in week_starts
        ]

    if request.GET.get("format") == "xlsx":
        workbook = build_status_summary_workbook(tasks, week_starts)
        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )
        response["Content-Disposition"] = (
            "attachment; filename="
            f'"status-summary-{current_week_start.isoformat()}.xlsx"'
        )
        workbook.save(response)
        return response

    grouping = get_status_summary_grouping(
        request.user,
        request.GET.get("grouping", ""),
    )
    groups = build_status_summary_groups(tasks, grouping)
    filled_count = sum(
        task.current_week_status_filled
        for task in tasks
    )

    departments = Department.objects.none()

    if role_code in {"head", "administrator"}:
        departments = Department.objects.filter(
            is_active=True,
        ).order_by("name")

    return render(
        request,
        "tracker/status_summary.html",
        {
            "groups": groups,
            "tasks_count": len(tasks),
            "filled_count": filled_count,
            "missing_count": len(tasks) - filled_count,
            "week_starts": week_starts,
            "current_week_start": current_week_start,
            "departments": departments,
            "selected_department": selected_department,
            "selected_grouping": grouping,
            "selected_search": search_value,
            "show_final": show_final,
            "final_only": final_only,
            "show_final_before_final_only": (
                show_final_before_final_only
            ),
            "missing_only": missing_only,
        },
    )


def check_task_access(user, task):
    role_code = user.role.code

    if role_code == "employee":
        if task.assignee_id != user.id:
            raise PermissionDenied

    elif role_code == "manager":
        if (
            not user.department_id
            or task.department_id != user.department_id
        ):
            raise PermissionDenied

    elif role_code not in {"head", "administrator"}:
        raise PermissionDenied


@login_required
def task_create(request):
    task = Task(created_by=request.user)

    if request.method == "POST":
        form = TaskForm(request.POST, instance=task, user=request.user)
        artifact_formset = TaskArtifactFormSet(
            request.POST,
            instance=task,
        )

        if form.is_valid() and artifact_formset.is_valid():
            with transaction.atomic():
                task = form.save(commit=False)
                task.created_by = request.user
                task.status = get_object_or_404(TaskStatus, code="new")
                task.save()

                artifacts = artifact_formset.save(commit=False)

                for artifact in artifacts:
                    artifact.created_by = request.user
                    artifact.save()

            return redirect("tracker:task-list")
    else:
        form = TaskForm(instance=task, user=request.user)
        artifact_formset = TaskArtifactFormSet(instance=task)

    return render(
        request,
        "tracker/task_form.html",
        {
            "form": form,
            "artifact_formset": artifact_formset,
        },
    )


@login_required
def task_update(request, task_id):
    task = get_object_or_404(Task, id=task_id)
    check_task_access(request.user, task)

    if request.method == "POST":
        form = TaskForm(request.POST, instance=task, user=request.user)
        artifact_formset = TaskArtifactFormSet(
            request.POST,
            instance=task,
        )

        if form.is_valid() and artifact_formset.is_valid():
            with transaction.atomic():
                form.save()

                artifacts = artifact_formset.save(commit=False)

                for deleted_artifact in artifact_formset.deleted_objects:
                    deleted_artifact.delete()

                for artifact in artifacts:
                    if not artifact.created_by_id:
                        artifact.created_by = request.user

                    artifact.save()

            return redirect("tracker:task-list")
    else:
        form = TaskForm(instance=task, user=request.user)
        artifact_formset = TaskArtifactFormSet(instance=task)

    return render(
        request,
        "tracker/task_form.html",
        {
            "form": form,
            "artifact_formset": artifact_formset,
            "task": task,
        },
    )


@login_required
@require_POST
def task_inline_update(request, task_id):
    task = get_object_or_404(
        Task.objects.select_related(
            "project_stream",
            "assignee",
            "department",
        ),
        id=task_id,
    )

    check_task_access(request.user, task)

    form = TaskInlineEditForm(
        request.POST,
        instance=task,
        prefix=f"task-edit-{task.id}",
        user=request.user,
    )

    if not form.is_valid():
        return JsonResponse(
            {
                "success": False,
                "errors": form.errors.get_json_data(),
            },
            status=400,
        )

    task = form.save()

    return JsonResponse(
        {
            "success": True,
            "task": {
                "id": task.id,
                "project_stream": {
                    "id": task.project_stream_id,
                    "name": task.project_stream.name,
                },
                "external_number": task.external_number or "",
                "external_url": task.external_url or "",
                "summary": task.summary,
                "department": {
                    "id": task.department_id,
                    "code": task.department.code,
                    "name": task.department.name,
                },
                "assignee": {
                    "id": task.assignee_id,
                    "name": task.assignee.name,
                },
            },
        }
    )


@login_required
@require_POST
def task_status_update(request, task_id):
    task = get_object_or_404(
        Task.objects.select_related("status"),
        id=task_id,
    )

    check_task_access(request.user, task)

    form = TaskStatusUpdateForm(
        request.POST,
        instance=task,
        prefix=f"task-status-{task.id}",
    )

    if not form.is_valid():
        return JsonResponse(
            {
                "success": False,
                "errors": form.errors.get_json_data(),
            },
            status=400,
        )

    task = form.save()

    return JsonResponse(
        {
            "success": True,
            "status": {
                "id": task.status_id,
                "name": task.status.name,
                "code": task.status.code,
                "is_final": task.status.is_final,
            },
        }
    )


@login_required
@require_POST
def task_artifact_create(request, task_id):
    task = get_object_or_404(Task, id=task_id)

    check_task_access(request.user, task)

    form = TaskArtifactForm(request.POST)

    if not form.is_valid():
        return JsonResponse(
            {
                "success": False,
                "errors": form.errors.get_json_data(),
            },
            status=400,
        )

    artifact = form.save(commit=False)
    artifact.task = task
    artifact.created_by = request.user
    artifact.save()

    return JsonResponse(
        {
            "success": True,
            "artifact": {
                "id": artifact.id,
                "name": artifact.name,
                "url": artifact.url,
            },
        },
        status=201,
    )

@login_required
@require_POST
def task_artifact_delete(request, task_id, artifact_id):
    task = get_object_or_404(Task, id=task_id)

    check_task_access(request.user, task)

    artifact = get_object_or_404(
        TaskArtifact,
        id=artifact_id,
        task=task,
    )

    artifact.delete()

    return JsonResponse(
        {
            "success": True,
            "artifact_id": artifact_id,
        }
    )
